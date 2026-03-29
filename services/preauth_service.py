import json
import re
import uuid
from datetime import datetime, timedelta
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from groq import Groq
import os
from .insurance_rules import check_rules_against_case

client = Groq(api_key=os.getenv("GROQ_API_KEY"))

# In-memory store for hackathon (pre-populated with 3 cases for presentation)
cases_store = [
    {
        "patient_id": "SIMPY-UID-001",
        "admission_id": "ADM-1001",
        "patient_name": "John Doe",
        "age": "45",
        "gender": "Male",
        "insurance_provider": "Star Health",
        "policy_number": "POL-77889",
        "tpa_name": "MediAssist",
        "diagnosis": "Acute Myocardial Infarction",
        "icd_code": "I21.9",
        "proposed_treatment": "Emergency Angioplasty",
        "completeness_score": 100,
        "risk_level": "High",
        "decision": "Review",
        "missing_fields": [],
        "suggestions": ["Verify cardiac biomarkers", "Check for prior stent history"],
        "reasoning": "High risk due to acute cardiac event",
        "created_at": (datetime.now() - timedelta(days=2)).isoformat(),
        "rule_check": {
            "total_rules": 32,
            "passed": 28,
            "failed": 2,
            "warnings": 2,
            "rule_score": 88,
            "results": []
        }
    },
    {
        "patient_id": "SIMPY-UID-002",
        "admission_id": "ADM-1002",
        "patient_name": "Jane Smith",
        "age": "32",
        "gender": "Female",
        "insurance_provider": "HDFC Ergo",
        "policy_number": "POL-11223",
        "tpa_name": "United Health",
        "diagnosis": "Appendicitis",
        "icd_code": "K35.8",
        "proposed_treatment": "Laparoscopic Appendectomy",
        "completeness_score": 100,
        "risk_level": "Low",
        "decision": "Approve",
        "missing_fields": [],
        "suggestions": ["Confirm NPO status", "Standard surgical protocols"],
        "reasoning": "Standard emergency procedure with no complications",
        "created_at": (datetime.now() - timedelta(days=1)).isoformat(),
        "rule_check": {
            "total_rules": 32,
            "passed": 31,
            "failed": 0,
            "warnings": 1,
            "rule_score": 97,
            "results": []
        }
    },
    {
        "patient_id": "SIMPY-UID-101",
        "admission_id": "ADM-2024",
        "patient_name": "Anita Verma",
        "age": "29",
        "gender": "Female",
        "insurance_provider": "Star Health",
        "policy_number": "SH-99011",
        "tpa_name": "MediAssist",
        "diagnosis": "Acute Appendicitis with Localized Peritonitis",
        "icd_code": "K35.3",
        "proposed_treatment": "Laparoscopic Appendectomy",
        "completeness_score": 92,
        "risk_level": "Low",
        "decision": "Approve",
        "missing_fields": [],
        "suggestions": [
            "Clinical justification matches ICD-10 criteria",
            "Ensure post-op recovery plan is uploaded after surgery",
            "Length of stay (3 days) is within standard limits"
        ],
        "reasoning": "High documentation completeness and definitive surgical indication.",
        "created_at": "2024-03-28T10:30:00Z"
    }
]
patient_counter = [102]
admission_counter = [2025]
enhancement_store = []

def generate_patient_id() -> str:
    pid = f"SIMPY-UID-{str(patient_counter[0]).zfill(3)}"
    patient_counter[0] += 1
    return pid

def generate_admission_id() -> str:
    aid = f"ADM-{str(admission_counter[0]).zfill(4)}"
    admission_counter[0] += 1
    return aid

def extract_pdf_text(pdf_path: str) -> str:
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text()
    return text

def validate_fields(patient_name: str, age: str, gender: str, insurance_provider: str, 
                    policy_number: str, tpa_name: str, pdf_text: str):
    required = {
        "patient_name": patient_name,
        "age": str(age),
        "gender": gender,
        "insurance_provider": insurance_provider,
        "policy_number": policy_number,
        "tpa_name": tpa_name,
    }
    missing = [k for k, v in required.items() if not v or (isinstance(v, str) and v.strip() == "")]
    
    # Check PDF for key medical fields
    pdf_fields = {
        "diagnosis": bool(re.search(r'diagno', pdf_text, re.I)),
        "icd_code": bool(re.search(r'[A-Z]\d{2}', pdf_text)),
        "proposed_treatment": bool(re.search(r'treatment|procedure|surgery', pdf_text, re.I)),
        "referral_letter": bool(re.search(r'referral|refer', pdf_text, re.I)),
    }
    
    for field, found in pdf_fields.items():
        if not found:
            missing.append(field)
    
    total_fields = len(required) + len(pdf_fields)
    filled = total_fields - len(missing)
    score = round((filled / total_fields) * 100)
    
    return missing, score

def analyze_with_groq(patient_name: str, age: str, gender: str, insurance_provider: str,
                       tpa_name: str, pdf_text: str, missing: list, score: int) -> dict:
    
    pdf_snippet = pdf_text[:1500] if pdf_text else ""
    
    prompt = f"""You are a senior healthcare pre-authorization auditor.

Patient Information:
- Name: {patient_name}
- Age: {age}
- Gender: {gender}
- Insurance Provider: {insurance_provider}
- TPA: {tpa_name}

Extracted PDF Content (first 1500 chars):
{pdf_snippet}

Validation Results:
- Completeness Score: {score}%
- Missing Fields: {', '.join(missing) if missing else 'None'}

Based on this pre-authorization request, analyze and return ONLY valid JSON:
{{
  "diagnosis": "extracted diagnosis or Unknown",
  "icd_code": "extracted ICD code like I50.3 or Unknown",
  "proposed_treatment": "extracted treatment or Unknown",
  "risk_level": "Low or Medium or High",
  "decision": "Approve or Review or Reject",
  "suggestions": ["suggestion 1", "suggestion 2", "suggestion 3"],
  "reasoning": "one line explanation of decision"
}}

Decision rules:
- Approve: score >= 80 and risk Low
- Review: score 50-79 or risk Medium or missing minor fields
- Reject: score < 50 or risk High or ICD code missing or diagnosis missing

Return ONLY the JSON object, no other text."""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.1,
        max_tokens=500
    )
    
    content = response.choices[0].message.content.strip()
    # Clean JSON
    content = re.sub(r'^```json\s*', '', content)
    content = re.sub(r'\s*```$', '', content)
    
    try:
        return json.loads(content)
    except:
        # Fallback if AI output is messy
        return {
            "diagnosis": "Review Required",
            "icd_code": "Pending",
            "proposed_treatment": "Incomplete",
            "risk_level": "Medium",
            "decision": "Review",
            "suggestions": ["AI analysis failed to parse. Please review manually."],
            "reasoning": "Parse error in AI response"
        }

def run_enhancement_audit(
    admission_id: str, patient_name: str, age: str, gender: str,
    insurance_provider: str, tpa_name: str, 
    original_approved_amt: float, enhancement_amt: float, 
    reason: str, pdf_path: str
) -> dict:
    pdf_text: str = extract_pdf_text(pdf_path)
    
    pdf_snippet = pdf_text[:1500] if pdf_text else ""
    
    prompt = f"""You are a senior TPA enhancement auditor.

Original Case: Admission ID {admission_id}
Patient: {patient_name}, {age}Y, {gender}
Insurance: {insurance_provider}, TPA: {tpa_name}
Original Approved Amount: Rs. {original_approved_amt}
Enhancement Requested: Rs. {enhancement_amt}
Reason for Enhancement: {reason}

Discharge Summary / Additional Docs (extracted):
{pdf_snippet}

Analyze this enhancement request and return ONLY valid JSON:
{{
  "enhancement_decision": "Approve or Review or Reject",
  "approved_enhancement_amt": "amount in numbers only",
  "copay_amount": "patient copay in numbers",
  "tpa_payable": "TPA payable amount in numbers",
  "deductions": "any deductions with reason",
  "risk_level": "Low or Medium or High",
  "suggestions": ["suggestion 1", "suggestion 2"],
  "reasoning": "one line explanation"
}}

Rules:
- Approve if documentation complete and amount justified
- Review if partial docs or borderline amount
- Reject if no clinical justification or excessive amount
- TPA payable = Enhancement amt - copay - deductions"""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.1,
        max_tokens=500
    )
    content = response.choices[0].message.content.strip()
    match = re.search(r'\{.*\}', content, re.DOTALL)
    if match:
        content = match.group(0)
    ai_result = json.loads(content)
    
    enhancement_report = {
        "admission_id": admission_id,
        "enhancement_id": f"ENH-{str(len(enhancement_store)+1).zfill(3)}",
        "patient_name": patient_name,
        "insurance_provider": insurance_provider,
        "tpa_name": tpa_name,
        "original_approved_amt": original_approved_amt,
        "enhancement_requested_amt": enhancement_amt,
        "enhancement_decision": ai_result.get("enhancement_decision"),
        "approved_enhancement_amt": ai_result.get("approved_enhancement_amt"),
        "copay_amount": ai_result.get("copay_amount"),
        "tpa_payable": ai_result.get("tpa_payable"),
        "deductions": ai_result.get("deductions"),
        "risk_level": ai_result.get("risk_level"),
        "suggestions": ai_result.get("suggestions", []),
        "reasoning": ai_result.get("reasoning"),
        "created_at": datetime.now().isoformat()
    }
    
    enhancement_store.append(enhancement_report)
    return enhancement_report

def auto_fill_preauth_form(pdf_path: str) -> dict:
    pdf_text: str = extract_pdf_text(pdf_path)
    
    pdf_snippet = pdf_text[:2000] if pdf_text else ""
    
    prompt = f"""You are a medical data extraction AI for healthcare pre-authorization.

Extract ALL patient and clinical information from this document.
Document content:
{pdf_snippet}

Return ONLY valid JSON with extracted data:
{{
  "patient_name": "full name or empty string",
  "age": "age as number or empty string",
  "gender": "Male or Female or Other or empty string",
  "insurance_provider": "insurance company name or empty string",
  "policy_number": "policy number or empty string",
  "tpa_name": "TPA name or empty string",
  "diagnosis": "primary diagnosis or empty string",
  "icd_code": "ICD-10 code or empty string",
  "proposed_treatment": "treatment or procedure or empty string",
  "doctor_name": "treating doctor or empty string",
  "hospital_name": "hospital name or empty string",
  "estimated_cost": "estimated cost in numbers or empty string",
  "confidence": "High or Medium or Low"
}}

Extract exactly what is written. Do not guess or hallucinate.
If a field is not found, return empty string for that field."""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.0,
        max_tokens=600
    )
    content = response.choices[0].message.content.strip()
    match = re.search(r'\{.*\}', content, re.DOTALL)
    if match:
        content = match.group(0)
    return json.loads(content)

def generate_excel(report: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pre-Auth Audit Report"
    
    # Header styling
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    headers = [
        "Patient ID", "Admission ID", "Patient Name", "Age", "Gender",
        "Diagnosis", "ICD Code", "Proposed Treatment", "Insurance Provider",
        "TPA Name", "Policy Number", "Completeness Score", "Risk Level",
        "Decision", "Missing Fields", "Suggestions", "Reasoning", "Date"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20
    
    # Decision color
    decision_colors = {
        "Approve": "00C851",
        "Review": "FF8800", 
        "Reject": "FF4444"
    }
    
    row_data = [
        report.get("patient_id"), report.get("admission_id"),
        report.get("patient_name"), report.get("age"), report.get("gender"),
        report.get("diagnosis"), report.get("icd_code"),
        report.get("proposed_treatment"), report.get("insurance_provider"),
        report.get("tpa_name"), report.get("policy_number"),
        f"{report.get('completeness_score')}%", report.get("risk_level"),
        report.get("decision"),
        ", ".join(report.get("missing_fields", [])),
        " | ".join(report.get("suggestions", [])),
        report.get("reasoning"),
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ]
    
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.alignment = Alignment(horizontal="left", wrap_text=True)
    
    # Color the decision cell
    decision = report.get("decision", "Review")
    decision_cell = ws.cell(row=2, column=14)
    color = decision_colors.get(decision, "FF8800")
    decision_cell.fill = PatternFill("solid", fgColor=color)
    decision_cell.font = Font(bold=True, color="FFFFFF")
    
    ws.row_dimensions[2].height = 40
    
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def run_preauth_audit(patient_name: str, age: str, gender: str, insurance_provider: str,
                       policy_number: str, tpa_name: str, pdf_path: str) -> dict:
    
    # Extract PDF
    pdf_text = extract_pdf_text(pdf_path)
    
    # Validate
    missing, score = validate_fields(
        patient_name, age, gender, 
        insurance_provider, policy_number, tpa_name, pdf_text
    )
    
    # AI Analysis
    ai_result = analyze_with_groq(
        patient_name, age, gender, insurance_provider,
        tpa_name, pdf_text, missing, score
    )
    
    # Generate IDs
    patient_id = generate_patient_id()
    admission_id = generate_admission_id()
    
    # Build report
    report = {
        "patient_id": patient_id,
        "admission_id": admission_id,
        "patient_name": patient_name,
        "age": age,
        "gender": gender,
        "insurance_provider": insurance_provider,
        "policy_number": policy_number,
        "tpa_name": tpa_name,
        "diagnosis": ai_result.get("diagnosis", "Unknown"),
        "icd_code": ai_result.get("icd_code", "Unknown"),
        "proposed_treatment": ai_result.get("proposed_treatment", "Unknown"),
        "completeness_score": score,
        "risk_level": ai_result.get("risk_level", "Medium"),
        "decision": ai_result.get("decision", "Review"),
        "missing_fields": missing,
        "suggestions": ai_result.get("suggestions", []),
        "reasoning": ai_result.get("reasoning", ""),
        "created_at": datetime.now().isoformat()
    }
    
    # Store in memory
    cases_store.append(report)
    
    # Rule Check
    rule_check = check_rules_against_case({**report, "pdf_text": pdf_text})
    report["rule_check"] = rule_check
    
    return report
